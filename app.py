from flask import Flask, render_template, request, redirect, session, send_file
from datetime import datetime, timezone, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill
import sqlite3
import json
import os

app = Flask(__name__)
app.secret_key = 'nautique_merville_secret'

MOT_DE_PASSE_BILAN = 'BNM2026'
MOT_DE_PASSE_BAR   = 'BNM2026BAR'

DB_PATH = os.path.join(os.path.dirname(__file__), 'bar.db')

bar_ferme = False

# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def maintenant():
    return datetime.now(timezone.utc) + timedelta(hours=2)

menu = {
    "Boissons": [
        {"nom": "Coca-Cola",    "prix": 2.50},
        {"nom": "Ice Tea",      "prix": 2.50},
        {"nom": "Oasis",        "prix": 2.50},
        {"nom": "Eau Minérale", "prix": 1.50}
    ],
    "Crêpes": [
        {"nom": "Sucre",      "prix": 3.00},
        {"nom": "Nutella",    "prix": 3.50},
        {"nom": "Confiture",  "prix": 3.50}
    ],
    "Glaces": [
        {"nom": "Magnum",         "prix": 3.00},
        {"nom": "Cornet Vanille", "prix": 2.50}
    ],
    "Planches": [
        {"nom": " Planche apéritif (saucisson,légume,jambon,cornichon,cacahuètes...)",  "prix":  8.00},
        {"nom": " Planche charcuterie (saucisson,pâté,jambon,fromage,cacahuètes...)",  "prix": 10.00},
        {"nom": " Planche fromage (brie,camembert,comté,pain,cacahuètes...)",           "prix":  8.50}
    ]
}

# ─────────────────────────────────────────────
# Base de données
# ─────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS commandes (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                produits    TEXT    NOT NULL,
                heure       TEXT    NOT NULL,
                statut      TEXT    NOT NULL DEFAULT "En preparation",
                total       REAL    NOT NULL,
                note        TEXT    DEFAULT "",
                heure_prete TEXT    DEFAULT NULL,
                archivee    INTEGER NOT NULL DEFAULT 0
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS config (
                cle    TEXT PRIMARY KEY,
                valeur TEXT NOT NULL
            )
        ''')
        # offset = id du dernier ticket avant remise à zéro
        # numero affiché = id - offset
        conn.execute('INSERT OR IGNORE INTO config (cle, valeur) VALUES ("numero_offset", "0")')
        conn.commit()

init_db()

def get_offset():
    with get_db() as conn:
        row = conn.execute('SELECT valeur FROM config WHERE cle = "numero_offset"').fetchone()
    return int(row['valeur']) if row else 0

def set_offset(val):
    with get_db() as conn:
        conn.execute('UPDATE config SET valeur = ? WHERE cle = "numero_offset"', (str(val),))
        conn.commit()

# ─────────────────────────────────────────────
# Conversion ligne DB -> dict
# Le numéro affiché = id - offset (repart de 1 après reset numérotation)
# L'id réel en DB est conservé pour les liens /suivi/, /recu/, etc.
# ─────────────────────────────────────────────

def row_to_dict(row, offset=0):
    d = dict(row)
    d['produits']      = json.loads(d['produits'])
    d['heure']         = datetime.fromisoformat(d['heure'])
    d['numero_affiche'] = d['id'] - offset
    if d['heure_prete']:
        d['heure_prete'] = datetime.fromisoformat(d['heure_prete'])
    else:
        d.pop('heure_prete', None)
    return d

def get_commandes_actives():
    offset = get_offset()
    with get_db() as conn:
        rows = conn.execute(
            'SELECT * FROM commandes WHERE archivee = 0 ORDER BY id'
        ).fetchall()
    return [row_to_dict(r, offset) for r in rows]

def get_historique():
    offset = get_offset()
    with get_db() as conn:
        rows = conn.execute(
            'SELECT * FROM commandes WHERE archivee = 1 ORDER BY heure'
        ).fetchall()
    return [row_to_dict(r, offset) for r in rows]

def get_commande_by_id(commande_id):
    offset = get_offset()
    with get_db() as conn:
        row = conn.execute(
            'SELECT * FROM commandes WHERE id = ?', (commande_id,)
        ).fetchone()
    return row_to_dict(row, offset) if row else None

def get_historique_du_jour(date_str):
    """Retourne les commandes archivées d'un jour donné (format YYYY-MM-DD)."""
    offset = get_offset()
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM commandes WHERE archivee = 1 AND heure LIKE ? ORDER BY heure",
            (date_str + '%',)
        ).fetchall()
    return [row_to_dict(r, offset) for r in rows]

# ─────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────

@app.route('/')
def afficher_menu():
    return render_template('menu.html', menu=menu, bar_ferme=bar_ferme)


@app.route('/commander', methods=['POST'])
def prendre_commande():
    if bar_ferme:
        return redirect('/')

    produits_choisis = request.form.getlist('produits')
    note = request.form.get('note', '').strip()

    if produits_choisis:
        total_commande = 0
        for prod in produits_choisis:
            for categorie in menu.values():
                for item in categorie:
                    if item['nom'] == prod or prod.endswith(item['nom']):
                        total_commande += item['prix']
                        break

        with get_db() as conn:
            cur = conn.execute(
                'INSERT INTO commandes (produits, heure, statut, total, note) VALUES (?, ?, ?, ?, ?)',
                (json.dumps(produits_choisis), maintenant().isoformat(),
                 "En préparation", total_commande, note)
            )
            conn.commit()
            new_id = cur.lastrowid

        return redirect(f'/suivi/{new_id}')

    return redirect('/')


@app.route('/commander-bar', methods=['POST'])
def commander_bar():
    if not session.get('bar_ok'):
        return redirect('/login_bar')

    produits_choisis = request.form.getlist('produits')
    note = request.form.get('note', '').strip()

    if produits_choisis:
        total_commande = 0
        for prod in produits_choisis:
            for categorie in menu.values():
                for item in categorie:
                    if item['nom'] == prod or prod.endswith(item['nom']):
                        total_commande += item['prix']
                        break

        with get_db() as conn:
            conn.execute(
                'INSERT INTO commandes (produits, heure, statut, total, note, archivee) VALUES (?, ?, ?, ?, ?, ?)',
                (json.dumps(produits_choisis), maintenant().isoformat(),
                 "Récupérée", total_commande, note, 1)
            )
            conn.commit()

    return redirect('/bar')


@app.route('/suivi/<int:commande_id>')
def suivi_commande(commande_id):
    commande = get_commande_by_id(commande_id)
    if commande:
        return render_template('suivi.html', commande=commande)
    return "Commande introuvable", 404


@app.route('/bar')
def ecran_bar():
    if not session.get('bar_ok'):
        return redirect('/login_bar')
    now = maintenant()
    commandes = get_commandes_actives()
    for c in commandes:
        c['attente'] = int((now - c['heure']).total_seconds() / 60)
    return render_template('bar.html', commandes=commandes, bar_ferme=bar_ferme, menu=menu)


@app.route('/login_bar', methods=['GET', 'POST'])
def login_bar():
    if request.method == 'POST':
        mdp = request.form.get('mot_de_passe', '')
        if mdp == MOT_DE_PASSE_BAR:
            session['bar_ok'] = True
            return redirect('/bar')
        else:
            return render_template('login_bar.html', erreur=True)
    return render_template('login_bar.html', erreur=False)


@app.route('/logout_bar')
def logout_bar():
    session.pop('bar_ok', None)
    return redirect('/login_bar')


@app.route('/toggle-bar', methods=['POST'])
def toggle_bar():
    global bar_ferme
    bar_ferme = not bar_ferme
    return redirect('/bar')


@app.route('/reset-numerotation', methods=['POST'])
def reset_numerotation():
    if not session.get('bar_ok'):
        return redirect('/login_bar')
    # L'offset devient l'id du dernier ticket inséré
    # => le prochain ticket affiché sera (max_id + 1) - (max_id) = 1
    with get_db() as conn:
        row = conn.execute('SELECT MAX(id) as max_id FROM commandes').fetchone()
        max_id = row['max_id'] or 0
    set_offset(max_id)
    return redirect('/bar')


@app.route('/prete/<int:commande_id>')
def commande_prete(commande_id):
    with get_db() as conn:
        conn.execute(
            'UPDATE commandes SET statut = ?, heure_prete = ? WHERE id = ?',
            ("Prête ! Passez au comptoir", maintenant().isoformat(), commande_id)
        )
        conn.commit()
    return redirect('/bar')


@app.route('/archiver/<int:commande_id>')
def archiver_commande(commande_id):
    with get_db() as conn:
        conn.execute(
            'UPDATE commandes SET statut = ?, archivee = 1 WHERE id = ?',
            ("Récupérée", commande_id)
        )
        conn.commit()
    return redirect('/bar')


@app.route('/archiver-tout', methods=['POST'])
def archiver_tout():
    with get_db() as conn:
        conn.execute(
            'UPDATE commandes SET statut = ?, archivee = 1 WHERE archivee = 0',
            ("Récupérée",)
        )
        conn.commit()
    return redirect('/bar')


@app.route('/bilan', methods=['GET', 'POST'])
def afficher_bilan():
    if request.method == 'POST' and 'mot_de_passe' in request.form:
        mdp = request.form.get('mot_de_passe', '')
        if mdp == MOT_DE_PASSE_BILAN:
            session['bilan_ok'] = True
        else:
            return render_template('login_bilan.html', erreur=True)

    if not session.get('bilan_ok'):
        return render_template('login_bilan.html', erreur=False)

    historique = get_historique()

    total_recettes   = sum(c['total'] for c in historique)
    total_articles   = 0
    stats_produits   = {}
    stats_categories = {}
    nb_commandes     = len(historique)
    panier_moyen     = round(total_recettes / nb_commandes, 2) if nb_commandes > 0 else 0

    commandes_par_heure = {}
    for c in historique:
        heure = c['heure'].strftime('%H:00')
        commandes_par_heure[heure] = commandes_par_heure.get(heure, 0) + 1

    heure_pointe = max(commandes_par_heure, key=commandes_par_heure.get) if commandes_par_heure else None

    for c in historique:
        for prod in c['produits']:
            total_articles += 1
            stats_produits[prod] = stats_produits.get(prod, 0) + 1
            for cat, items in menu.items():
                for item in items:
                    if item['nom'] == prod or prod.endswith(item['nom']):
                        stats_categories[cat] = stats_categories.get(cat, 0) + 1
                        break

    produit_star = max(stats_produits, key=stats_produits.get) if stats_produits else None
    commandes_par_heure_triees = dict(sorted(commandes_par_heure.items()))

    temps_prep_list = []
    for c in historique:
        if 'heure_prete' in c:
            duree = (c['heure_prete'] - c['heure']).total_seconds() / 60
            temps_prep_list.append(duree)
    temps_moyen_prep = round(sum(temps_prep_list) / len(temps_prep_list), 1) if temps_prep_list else None

    # ── Comparaison J vs J-1 ──
    today_str     = maintenant().strftime('%Y-%m-%d')
    yesterday_str = (maintenant() - timedelta(days=1)).strftime('%Y-%m-%d')

    hist_j   = get_historique_du_jour(today_str)
    hist_j1  = get_historique_du_jour(yesterday_str)

    def stats_jour(hist):
        if not hist:
            return None
        ca      = round(sum(c['total'] for c in hist), 2)
        nb      = len(hist)
        articles = sum(len(c['produits']) for c in hist)
        panier   = round(ca / nb, 2) if nb else 0
        return {'ca': ca, 'nb': nb, 'articles': articles, 'panier': panier}

    compa_j  = stats_jour(hist_j)
    compa_j1 = stats_jour(hist_j1)

    def evolution(val_j, val_j1):
        """Retourne le % d'évolution arrondi, ou None si pas de J-1."""
        if val_j1 is None or val_j1 == 0:
            return None
        return round((val_j - val_j1) / val_j1 * 100, 1)

    compa_evol = None
    if compa_j and compa_j1:
        compa_evol = {
            'ca':      evolution(compa_j['ca'],      compa_j1['ca']),
            'nb':      evolution(compa_j['nb'],       compa_j1['nb']),
            'articles': evolution(compa_j['articles'], compa_j1['articles']),
            'panier':  evolution(compa_j['panier'],   compa_j1['panier']),
        }

    return render_template('bilan.html',
                           historique=historique,
                           total_recettes=total_recettes,
                           total_articles=total_articles,
                           stats_produits=stats_produits,
                           stats_categories=stats_categories,
                           panier_moyen=panier_moyen,
                           heure_pointe=heure_pointe,
                           nb_commandes=nb_commandes,
                           produit_star=produit_star,
                           commandes_par_heure=commandes_par_heure_triees,
                           now=maintenant(),
                           menu=menu,
                           temps_moyen_prep=temps_moyen_prep,
                           compa_j=compa_j,
                           compa_j1=compa_j1,
                           compa_evol=compa_evol,
                           today_str=today_str,
                           yesterday_str=yesterday_str)


@app.route('/export-excel')
def export_excel():
    if not session.get('bilan_ok'):
        return redirect('/bilan')

    historique = get_historique()

    wb = openpyxl.Workbook()
    titre_font  = Font(bold=True, size=13, color="FFFFFF")
    titre_fill  = PatternFill("solid", fgColor="0056B3")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")

    # ── Onglet 1 : Résumé ──
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.append(["Bilan du " + maintenant().strftime('%d/%m/%Y')])
    ws1['A1'].font = titre_font
    ws1['A1'].fill = titre_fill
    ws1.append([])
    ws1.append(["Indicateur", "Valeur"])
    for cell in ws1[3]:
        cell.font = header_font
        cell.fill = header_fill

    nb_commandes    = len(historique)
    total_recettes  = sum(c['total'] for c in historique)
    panier_moyen    = round(total_recettes / nb_commandes, 2) if nb_commandes > 0 else 0
    stats_produits  = {}
    commandes_par_heure = {}
    for c in historique:
        heure = c['heure'].strftime('%H:00')
        commandes_par_heure[heure] = commandes_par_heure.get(heure, 0) + 1
        for prod in c['produits']:
            stats_produits[prod] = stats_produits.get(prod, 0) + 1

    produit_star = max(stats_produits, key=stats_produits.get) if stats_produits else "—"
    heure_pointe = max(commandes_par_heure, key=commandes_par_heure.get) if commandes_par_heure else "—"

    ws1.append(["Nombre de commandes",    nb_commandes])
    ws1.append(["Chiffre d'affaires (€)", total_recettes])
    ws1.append(["Panier moyen (€)",       panier_moyen])
    ws1.append(["Produit star",           produit_star])
    ws1.append(["Heure de pointe",        heure_pointe])
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 20

    # ── Onglet 2 : Ventes par produit ──
    ws2 = wb.create_sheet("Ventes par produit")
    ws2.append(["Produit", "Quantité", "Recette (€)"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    for prod, qte in sorted(stats_produits.items(), key=lambda x: x[1], reverse=True):
        prix_unit = 0
        for cat in menu.values():
            for item in cat:
                if item['nom'] == prod or prod.endswith(item['nom']):
                    prix_unit = item['prix']
        ws2.append([prod, qte, round(qte * prix_unit, 2)])
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 15

    # ── Onglet 3 : Détail commandes ──
    ws3 = wb.create_sheet("Détail commandes")
    ws3.append(["#", "Heure", "Produits", "Note", "Total (€)"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
    for c in sorted(historique, key=lambda x: x['heure']):
        ws3.append([
            f"#{c['numero_affiche']:03d}",
            c['heure'].strftime('%H:%M'),
            ", ".join(c['produits']),
            c.get('note', ''),
            c['total']
        ])
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 50
    ws3.column_dimensions['D'].width = 20
    ws3.column_dimensions['E'].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"bilan_{maintenant().strftime('%d-%m-%Y')}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route('/reset', methods=['POST'])
def reset_bilan():
    with get_db() as conn:
        conn.execute('DELETE FROM commandes WHERE archivee = 1')
        conn.commit()
    return redirect('/bilan')


@app.route('/logout_bilan')
def logout_bilan():
    session.pop('bilan_ok', None)
    return redirect('/bilan')


@app.route('/recu/<int:commande_id>')
def recu_commande(commande_id):
    commande = get_commande_by_id(commande_id)
    if commande:
        return render_template('recu.html', commande=commande, menu=menu)
    return "Commande introuvable", 404


if __name__ == '__main__':
    app.run(debug=True)
